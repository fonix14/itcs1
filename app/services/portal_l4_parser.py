from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


@dataclass
class PortalL4ParseResult:
    tasks: List[Dict[str, Any]]
    invalid: int
    total: int
    headers: List[str]  # debug


def _s(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M")
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer():
            v = int(v)
        return str(v)
    return str(v).strip()


def _parse_dt_str(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    for fmt in (
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d-%m-%Y",
        "%d.%m.%Y",
        "%Y-%m-%d",
    ):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            pass
    return ""


def _norm(s: str) -> str:
    return (s or "").strip().lower().replace("\n", " ").replace("\t", " ")


def _find_col_by_any(header: Dict[str, int], *, exact: List[str], contains: List[str]) -> Optional[int]:
    # 1) exact match
    for name in exact:
        if name in header:
            return header[name]
    # 2) normalized exact
    norm_map = {_norm(k).replace(" ", "").replace(",", ""): v for k, v in header.items()}
    for name in exact:
        key = _norm(name).replace(" ", "").replace(",", "")
        if key in norm_map:
            return norm_map[key]
    # 3) contains keywords
    for k, v in header.items():
        kl = _norm(k)
        for part in contains:
            if part in kl:
                return v
    return None


def parse_portal_l4_xlsx(file_bytes: bytes) -> PortalL4ParseResult:
    wb = openpyxl.load_workbook(filename=bytes_to_temp(file_bytes), data_only=True)
    ws = wb["Tasks"] if "Tasks" in wb.sheetnames else wb[wb.sheetnames[0]]

    header: Dict[str, int] = {}
    headers_list: List[str] = []
    for col in range(1, ws.max_column + 1):
        name = ws.cell(1, col).value
        if name is None:
            continue
        key = str(name).strip()
        header[key] = col
        headers_list.append(key)

    req = [
        "Идентификатор Портала",
        "Дата создания",
        "Уровень 4",
        "Текст обращения",
        "Номер магазина",
    ]
    missing = [x for x in req if x not in header]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    # robust columns
    sla_col = _find_col_by_any(
        header,
        exact=["Дата SLA", "Контроль до", "SLA"],
        contains=["sla", "контрол", "крайн", "дедлайн", "срок"],
    )
    comments_col = _find_col_by_any(
        header,
        exact=["Комментарии", "Комментарий", "Коммент"],
        contains=["коммент", "примеч", "описан", "ответ", "результат"],
    )
    loc_col = _find_col_by_any(
        header,
        exact=["Местонахождение", "Местоположение", "Адрес", "Локация"],
        contains=["место", "локац", "адрес", "тц", "магазин находится"],
    )

    tasks: List[Dict[str, Any]] = []
    invalid = 0
    total = max(0, ws.max_row - 1)

    for r in range(2, ws.max_row + 1):
        portal_task_id = _s(ws.cell(r, header["Идентификатор Портала"]).value)
        created_at = _s(ws.cell(r, header["Дата создания"]).value)
        level4 = _s(ws.cell(r, header["Уровень 4"]).value)
        text = _s(ws.cell(r, header["Текст обращения"]).value)
        store_no = _s(ws.cell(r, header["Номер магазина"]).value)

        comments = ""
        if comments_col is not None:
            comments = _s(ws.cell(r, comments_col).value)

        location = ""
        if loc_col is not None:
            location = _s(ws.cell(r, loc_col).value)

        sla_date = ""
        if sla_col is not None:
            raw = ws.cell(r, sla_col).value
            if isinstance(raw, datetime):
                sla_date = _s(raw)
            else:
                sraw = _s(raw)
                sla_date = _parse_dt_str(sraw) or sraw

        if not portal_task_id or not store_no:
            invalid += 1
            continue

        tasks.append(
            {
                "portal_task_id": portal_task_id,
                "created_at": created_at or "—",
                "level4": level4 or "—",
                "text": text or "—",
                "comments": comments or "—",
                "store_no": store_no or "—",
                "sla_date": sla_date or "—",
                "location": location or "—",
            }
        )

    return PortalL4ParseResult(tasks=tasks, invalid=invalid, total=total, headers=headers_list)


# --- temp helper ---
import tempfile, os

def bytes_to_temp(b: bytes) -> str:
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    with open(path, "wb") as f:
        f.write(b)
    return path

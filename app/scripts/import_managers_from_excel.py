from __future__ import annotations

import asyncio
import secrets
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

from app.auth import make_password_pair
from app.db import SessionLocal


EMAIL_ALIASES = {
    "email", "почта", "e-mail", "email менеджера", "почта менеджера"
}
NAME_ALIASES = {
    "full_name", "name", "фио", "менеджер", "ответственный", "имя"
}
STORE_ALIASES = {
    "store_no", "store", "номер магазина", "магазин", "store number"
}


def norm(v: object) -> str:
    return str(v or "").strip().lower()


def detect_columns(header: list[str]) -> dict[str, int]:
    result = {}
    for idx, col in enumerate(header):
        c = norm(col)
        if c in EMAIL_ALIASES and "email" not in result:
            result["email"] = idx
        elif c in NAME_ALIASES and "full_name" not in result:
            result["full_name"] = idx
        elif c in STORE_ALIASES and "store_no" not in result:
            result["store_no"] = idx
    return result


async def upsert_manager(session, full_name: str, email: str, password: str | None) -> str:
    res = await session.execute(
        text(
            """
            select id::text as id
            from users
            where lower(email) = lower(:email)
            limit 1
            """
        ),
        {"email": email},
    )
    row = res.mappings().first()

    salt = None
    password_hash = None
    if password:
        salt, password_hash = make_password_pair(password)

    if row:
        manager_id = row["id"]
        if password:
            await session.execute(
                text(
                    """
                    update users
                    set
                        full_name = :full_name,
                        is_active = true,
                        role = 'manager',
                        password_salt = :salt,
                        password_hash = :password_hash
                    where id = cast(:id as uuid)
                    """
                ),
                {
                    "id": manager_id,
                    "full_name": full_name,
                    "salt": salt,
                    "password_hash": password_hash,
                },
            )
        else:
            await session.execute(
                text(
                    """
                    update users
                    set
                        full_name = :full_name,
                        is_active = true,
                        role = 'manager'
                    where id = cast(:id as uuid)
                    """
                ),
                {
                    "id": manager_id,
                    "full_name": full_name,
                },
            )
        return manager_id

    if password:
        res = await session.execute(
            text(
                """
                insert into users (
                    id, full_name, email, role, is_active, password_salt, password_hash
                )
                values (
                    gen_random_uuid(), :full_name, :email, 'manager', true, :salt, :password_hash
                )
                returning id::text as id
                """
            ),
            {
                "full_name": full_name,
                "email": email,
                "salt": salt,
                "password_hash": password_hash,
            },
        )
    else:
        res = await session.execute(
            text(
                """
                insert into users (
                    id, full_name, email, role, is_active
                )
                values (
                    gen_random_uuid(), :full_name, :email, 'manager', true
                )
                returning id::text as id
                """
            ),
            {
                "full_name": full_name,
                "email": email,
            },
        )
    return res.mappings().first()["id"]


async def assign_store(session, store_no: str, manager_id: str) -> bool:
    res = await session.execute(
        text(
            """
            update stores
            set assigned_user_id = cast(:manager_id as uuid)
            where trim(store_no) = trim(:store_no)
            """
        ),
        {
            "manager_id": manager_id,
            "store_no": str(store_no).strip(),
        },
    )
    return (res.rowcount or 0) > 0


async def main():
    path = Path(sys.argv[1] if len(sys.argv) > 1 else "/mnt/data/менеджеры.xlsx")
    mode = (sys.argv[2] if len(sys.argv) > 2 else "generate").strip().lower()

    if not path.exists():
        print(f"ERROR: file not found: {path}")
        raise SystemExit(1)

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("ERROR: empty workbook")
        raise SystemExit(1)

    header = [str(x or "").strip() for x in rows[0]]
    cols = detect_columns(header)

    if "email" not in cols or "full_name" not in cols:
        print("ERROR: required columns not found")
        print("HEADER:", header)
        raise SystemExit(1)

    created = 0
    updated = 0
    assigned = 0
    passwords = []

    async with SessionLocal() as session:
        try:
            for raw in rows[1:]:
                if not raw:
                    continue

                email = str(raw[cols["email"]] or "").strip().lower()
                full_name = str(raw[cols["full_name"]] or "").strip()
                store_no = None
                if "store_no" in cols:
                    store_no = str(raw[cols["store_no"]] or "").strip()

                if not email or not full_name:
                    continue

                password = None
                if mode == "generate":
                    password = secrets.token_urlsafe(8)

                exists = await session.execute(
                    text("select 1 from users where lower(email)=lower(:email) limit 1"),
                    {"email": email},
                )
                existed = bool(exists.scalar())

                manager_id = await upsert_manager(
                    session=session,
                    full_name=full_name,
                    email=email,
                    password=password,
                )

                if existed:
                    updated += 1
                else:
                    created += 1

                if store_no:
                    ok = await assign_store(session, store_no, manager_id)
                    if ok:
                        assigned += 1

                if password:
                    passwords.append((full_name, email, password))

            await session.commit()

        except Exception as e:
            await session.rollback()
            print(f"ERROR: {e}")
            raise

    print(f"OK: created={created}, updated={updated}, assigned={assigned}")
    if passwords:
        print("\nGENERATED PASSWORDS:")
        for full_name, email, password in passwords:
            print(f"{full_name} | {email} | {password}")


if __name__ == "__main__":
    asyncio.run(main())

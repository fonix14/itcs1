from __future__ import annotations

import re
import uuid
from pathlib import Path
from collections import OrderedDict

import openpyxl

XLSX_PATH = Path("/mnt/data/менеджеры.xlsx")
SQL_OUT = Path("/opt/itcs/itcs_mvp_stage4/backend/sql/import_managers_from_excel.generated.sql")


def sql_q(value: str) -> str:
    return value.replace("'", "''")


def normalize_store_no(value) -> str:
    s = str(value).strip()
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def manager_uuid(name: str) -> str:
    return str(uuid.uuid5(uuid.NAMESPACE_DNS, f"itcs-manager::{name.strip()}"))


def store_uuid(store_no: str) -> str:
    return str(uuid.uuid5(uuid.NAMESPACE_DNS, f"itcs-store::{store_no.strip()}"))


def manager_email(name: str) -> str:
    slug = uuid.uuid5(uuid.NAMESPACE_DNS, f"itcs-manager-email::{name.strip()}").hex[:12]
    return f"manager-{slug}@local"


def main() -> None:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Файл не найден: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("Excel пустой")

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    if len(header) < 2:
        raise RuntimeError(f"Ожидались минимум 2 колонки, получено: {header}")

    col_store = None
    col_manager = None

    for i, name in enumerate(header):
        if name in ("№ магазина", "Номер магазина", "Магазин", "store_no"):
            col_store = i
        if name in ("Менеджер", "manager", "Ответственный"):
            col_manager = i

    if col_store is None or col_manager is None:
        raise RuntimeError(f"Не найдены нужные колонки. Header={header}")

    managers: OrderedDict[str, dict] = OrderedDict()
    store_map: OrderedDict[str, str] = OrderedDict()

    skipped = 0

    for idx, row in enumerate(rows[1:], start=2):
        store_raw = row[col_store] if col_store < len(row) else None
        manager_raw = row[col_manager] if col_manager < len(row) else None

        if store_raw is None and manager_raw is None:
            continue

        if store_raw is None or manager_raw is None:
            skipped += 1
            print(f"SKIP row {idx}: incomplete row -> {row}")
            continue

        store_no = normalize_store_no(store_raw)
        manager_name = str(manager_raw).strip()

        if not store_no or not manager_name:
            skipped += 1
            print(f"SKIP row {idx}: blank values -> {row}")
            continue

        managers.setdefault(
            manager_name,
            {
                "id": manager_uuid(manager_name),
                "email": manager_email(manager_name),
                "full_name": manager_name,
                "display_name": manager_name,
            },
        )

        store_map[store_no] = manager_name

    if not managers:
        raise RuntimeError("Не найдено ни одного менеджера в Excel")

    sql_lines: list[str] = []
    sql_lines.append("BEGIN;")
    sql_lines.append("")

    sql_lines.append("-- ensure display_name exists")
    sql_lines.append("ALTER TABLE users ADD COLUMN IF NOT EXISTS display_name varchar(255);")
    sql_lines.append("")

    sql_lines.append("-- upsert managers")
    for manager_name, meta in managers.items():
        sql_lines.append(
            f"""INSERT INTO users (id, email, role, full_name, display_name, is_active)
VALUES (
    '{meta["id"]}'::uuid,
    '{sql_q(meta["email"])}',
    'manager',
    '{sql_q(meta["full_name"])}',
    '{sql_q(meta["display_name"])}',
    true
)
ON CONFLICT (email) DO UPDATE
SET
    full_name = EXCLUDED.full_name,
    display_name = EXCLUDED.display_name,
    is_active = true;"""
        )
        sql_lines.append("")

    sql_lines.append("-- ensure stores exist")
    for store_no, manager_name in store_map.items():
        meta = managers[manager_name]
        sql_lines.append(
            f"""INSERT INTO stores (id, store_no, name, assigned_user_id)
VALUES (
    '{store_uuid(store_no)}'::uuid,
    '{sql_q(store_no)}',
    'Store {sql_q(store_no)}',
    '{meta["id"]}'::uuid
)
ON CONFLICT (store_no) DO NOTHING;"""
        )
        sql_lines.append("")

    sql_lines.append("-- assign stores")
    for store_no, manager_name in store_map.items():
        meta = managers[manager_name]
        sql_lines.append(
            f"""UPDATE stores
SET
    name = COALESCE(NULLIF(name, ''), 'Store {sql_q(store_no)}'),
    assigned_user_id = '{meta["id"]}'::uuid
WHERE store_no = '{sql_q(store_no)}';"""
        )
        sql_lines.append("")

    sql_lines.append("-- disable temporary bootstrap manager")
    sql_lines.append(
        """UPDATE users
SET is_active = false
WHERE email = 'manager1@local';"""
    )
    sql_lines.append("")

    sql_lines.append("COMMIT;")
    sql_lines.append("")

    SQL_OUT.write_text("\n".join(sql_lines), encoding="utf-8")

    print(f"OK: SQL generated -> {SQL_OUT}")
    print(f"Managers found: {len(managers)}")
    print(f"Stores found: {len(store_map)}")
    print(f"Skipped rows: {skipped}")
    print("Managers:")
    for name in managers.keys():
        print(" -", name)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}")
        raise

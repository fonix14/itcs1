from __future__ import annotations

import hashlib
import io
from datetime import datetime, timedelta
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Upload, TaskStaging, ImportError, Store, Task, UploadMetrics,
    Anomaly, AnomalySeverity, AnomalyStatus
)

STATUS_WHITELIST = {"new", "in_progress", "done", "closed", "open"}
SLA_HOURS = {"critical": 24, "major": 72, "minor": 24 * 7}


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _due(sev: AnomalySeverity) -> datetime:
    return datetime.utcnow() + timedelta(hours=SLA_HOURS[sev.value])


def _severity_for(anomaly_type: str) -> AnomalySeverity:
    if anomaly_type in ("UNKNOWN_SHEET", "COVERAGE_DROP"):
        return AnomalySeverity.major
    if anomaly_type in ("MISSING_MANAGER", "STORE_CHANGED", "CONFLICTING_ROWS"):
        return AnomalySeverity.major
    if anomaly_type in ("INVALID_RATIO_HIGH",):
        return AnomalySeverity.minor
    return AnomalySeverity.minor


async def create_anomaly(
    db: AsyncSession,
    anomaly_type: str,
    details: dict[str, Any],
    related_upload_id=None,
    related_task_id=None,
) -> None:
    sev = _severity_for(anomaly_type)
    db.add(
        Anomaly(
            anomaly_type=anomaly_type,
            severity=sev,
            status=AnomalyStatus.open,
            related_upload_id=related_upload_id,
            related_task_id=related_task_id,
            details=details,
            due_at=_due(sev),
        )
    )


def _norm(h: Any) -> str:
    return str(h).strip().lower().replace(" ", "_")


def _pick_sheet(wb):
    if "tasks" in wb.sheetnames:
        return wb["tasks"]
    return wb[wb.sheetnames[0]]


def _parse_rows(file_bytes: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = _pick_sheet(wb)
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return [], ["EMPTY_SHEET"]

    header = [_norm(x) for x in rows[0]]
    required = {"portal_task_id", "store_no", "status"}
    if not required.issubset(set(header)):
        return [], ["UNKNOWN_SHEET"]

    idx = {h: i for i, h in enumerate(header) if h}
    out: list[dict[str, Any]] = []
    for rn, r in enumerate(rows[1:], start=2):  # excel row number
        def get(col: str):
            i = idx.get(col)
            return None if i is None else r[i]
        out.append(
            {
                "row_number": rn,
                "portal_task_id": (str(get("portal_task_id")).strip() if get("portal_task_id") is not None else None),
                "store_no": (str(get("store_no")).strip() if get("store_no") is not None else None),
                "status": (str(get("status")).strip().lower() if get("status") is not None else None),
                "sla": (str(get("sla")).strip().lower() if get("sla") is not None else None),
                "raw": {h: (r[idx[h]] if h in idx else None) for h in idx.keys()},
            }
        )
    return out, []


async def _baseline_seen(db: AsyncSession, profile_id: str) -> float | None:
    res = await db.execute(
        select(Upload.seen_tasks_count)
        .where(Upload.profile_id == profile_id)
        .order_by(Upload.uploaded_at.desc())
        .limit(3)
    )
    vals = [v for (v,) in res.all()]
    if not vals:
        return None
    return sum(vals) / len(vals)


def _coverage_drop(seen: int, baseline: float | None) -> tuple[bool, int | None, float | None]:
    if baseline is None or baseline <= 0:
        return False, None, None
    abs_drop = int(round(baseline - seen))
    rel_drop = (baseline - seen) / baseline if baseline else 0.0

    # thresholds
    if baseline < 10:
        abs_threshold = 3
        rel_threshold = 0.50
    else:
        abs_threshold = max(5, int(0.15 * baseline))
        rel_threshold = 0.30

    drop = (abs_drop >= abs_threshold) and (rel_drop >= rel_threshold)
    return drop, abs_drop, rel_drop


async def compute_trust(db: AsyncSession) -> dict[str, Any]:
    reasons: list[str] = []
    trust = "GREEN"

    last = await db.execute(select(func.max(Upload.uploaded_at)))
    last_import_at = last.scalar_one_or_none()

    if last_import_at is None:
        trust = "RED"
        reasons.append("no_import_yet")
    else:
        hours = (datetime.utcnow() - last_import_at.replace(tzinfo=None)).total_seconds() / 3600.0
        if hours > 48:
            trust = "RED"
            reasons.append("no_import_over_48h")

    # open critical anomalies -> RED
    crit = await db.execute(
        select(func.count(Anomaly.id))
        .where(Anomaly.status == AnomalyStatus.open)
        .where(Anomaly.severity == AnomalySeverity.critical)
    )
    if (crit.scalar_one() or 0) > 0:
        trust = "RED"
        reasons.append("critical_anomaly_open")

    # coverage drop on latest upload -> YELLOW (unless already RED)
    cov = await db.execute(
        select(UploadMetrics.coverage_drop, Upload.invalid_ratio)
        .join(Upload, Upload.id == UploadMetrics.upload_id)
        .order_by(Upload.uploaded_at.desc())
        .limit(1)
    )
    row = cov.first()
    coverage_drop = None
    invalid_ratio = None
    if row:
        coverage_drop = bool(row[0])
        invalid_ratio = float(row[1])
        if trust != "RED" and coverage_drop:
            trust = "YELLOW"
            reasons.append("coverage_drop")
        if trust != "RED" and invalid_ratio > 0.20:
            trust = "YELLOW"
            reasons.append("invalid_ratio_over_20")

    pending = await db.execute(select(func.count(Anomaly.id)).where(Anomaly.status == AnomalyStatus.open))
    return {
        "trust_level": trust,
        "reasons": reasons,
        "last_import_at": last_import_at,
        "invalid_ratio": invalid_ratio,
        "coverage_drop": coverage_drop,
        "pending_anomalies": int(pending.scalar_one() or 0),
    }


async def import_xlsx_soft(db: AsyncSession, file_name: str, file_bytes: bytes, profile_id: str) -> dict[str, Any]:
    file_hash = sha256_bytes(file_bytes)

    # idempotency
    existing = await db.execute(select(Upload).where(Upload.file_hash == file_hash))
    ex = existing.scalar_one_or_none()
    if ex:
        trust = await compute_trust(db)
        return {"upload": ex, "anomalies_created": 0, "idempotent": True, "trust": trust}

    parsed_rows, parse_errs = _parse_rows(file_bytes)
    upload = Upload(file_name=file_name, file_hash=file_hash, profile_id=profile_id)
    db.add(upload)
    await db.flush()

    anomalies_created = 0
    created_tasks = 0
    updated_tasks = 0
    if parse_errs:
        await create_anomaly(db, "UNKNOWN_SHEET", {"errors": parse_errs}, related_upload_id=upload.id)
        anomalies_created += 1
        upload.total_rows = 0
        upload.valid_rows = 0
        upload.invalid_rows = 0
        upload.invalid_ratio = 1.0
        upload.seen_tasks_count = 0
        # still write metrics row
        db.add(UploadMetrics(upload_id=upload.id, baseline_seen=None, abs_drop=None, rel_drop=None, coverage_drop=False))
        await db.commit()
        trust = await compute_trust(db)
        return {"upload": upload, "anomalies_created": anomalies_created, "idempotent": False, "trust": trust, "created_tasks": created_tasks, "updated_tasks": updated_tasks}

    upload.total_rows = len(parsed_rows)

    seen_portal_ids: set[str] = set()
    store_map: dict[str, Store] = {}
    # prefetch stores by store_no present
    store_nos = sorted({r["store_no"] for r in parsed_rows if r.get("store_no")})
    if store_nos:
        res = await db.execute(select(Store).where(Store.store_no.in_(store_nos)))
        for s in res.scalars().all():
            store_map[s.store_no] = s

    # detect conflicting portal_task_id mapping to different store_no in same upload
    by_pt: dict[str, set[str]] = {}
    for r in parsed_rows:
        pt = r.get("portal_task_id")
        sn = r.get("store_no")
        if pt and sn:
            by_pt.setdefault(pt, set()).add(sn)
    conflicted = {pt: sorted(list(sns)) for pt, sns in by_pt.items() if len(sns) > 1}
    if conflicted:
        await create_anomaly(db, "CONFLICTING_ROWS", {"conflicts": conflicted}, related_upload_id=upload.id)
        anomalies_created += 1

    valid_rows = 0
    invalid_rows = 0

    # stage + validate
    for r in parsed_rows:
        err = None
        pt = r.get("portal_task_id")
        sn = r.get("store_no")
        st = r.get("status")
        sla = r.get("sla")

        if not pt:
            err = "MISSING_PORTAL_TASK_ID"
        elif not sn:
            err = "MISSING_STORE_NO"
        elif st not in STATUS_WHITELIST:
            err = "INVALID_STATUS"
        elif sn not in store_map:
            err = "STORE_NOT_FOUND"
        else:
            store = store_map[sn]
            if store.assigned_user_id is None:
                # dispatcher should assign later
                await create_anomaly(db, "MISSING_MANAGER", {"store_no": sn}, related_upload_id=upload.id)
                anomalies_created += 1

        db.add(TaskStaging(
            upload_id=upload.id,
            row_number=int(r["row_number"]),
            portal_task_id=pt,
            store_no=sn,
            status=st,
            sla=sla,
            raw_row=r.get("raw") or {},
            validation_error=err,
        ))

        if err:
            invalid_rows += 1
            db.add(ImportError(upload_id=upload.id, row_number=int(r["row_number"]), error_code=err, details={"row": r.get("raw") or {}}))
        else:
            valid_rows += 1
            seen_portal_ids.add(pt)

    upload.valid_rows = valid_rows
    upload.invalid_rows = invalid_rows
    upload.invalid_ratio = (invalid_rows / upload.total_rows) if upload.total_rows else 0.0
    upload.seen_tasks_count = len(seen_portal_ids)

    # commit valid tasks (upsert-ish)
    now = datetime.utcnow()
    for pt in seen_portal_ids:
        # pick first staging row for pt
        r = next(x for x in parsed_rows if x.get("portal_task_id") == pt)
        sn = r["store_no"]
        store = store_map[sn]
        status = r["status"]
        sla = r.get("sla")

        q = await db.execute(select(Task).where(Task.portal_task_id == pt))
        task = q.scalar_one_or_none()
        if task is None:
            task = Task(portal_task_id=pt, store_id=store.id, status=status, sla=sla, last_seen_at=now, created_at=now)
            db.add(task)
            created_tasks += 1
        else:
            changed = (task.status != status) or (task.sla != sla) or (task.store_id != store.id)
            if changed:
                updated_tasks += 1

            # store changed anomaly
            if task.store_id != store.id:
                await create_anomaly(db, "STORE_CHANGED", {"portal_task_id": pt, "from_store_id": str(task.store_id), "to_store_id": str(store.id)}, related_upload_id=upload.id, related_task_id=task.id)
                anomalies_created += 1
                task.store_id = store.id
            task.status = status
            task.sla = sla
            task.last_seen_at = now

    # metrics baseline/drop
    baseline = await _baseline_seen(db, profile_id)
    drop, abs_drop, rel_drop = _coverage_drop(upload.seen_tasks_count, baseline)
    db.add(UploadMetrics(upload_id=upload.id, baseline_seen=baseline, abs_drop=abs_drop, rel_drop=rel_drop, coverage_drop=drop))

    if upload.invalid_ratio > 0.20:
        await create_anomaly(db, "INVALID_RATIO_HIGH", {"invalid_ratio": upload.invalid_ratio}, related_upload_id=upload.id)
        anomalies_created += 1

    if drop:
        await create_anomaly(db, "COVERAGE_DROP", {"baseline": baseline, "seen": upload.seen_tasks_count, "abs_drop": abs_drop, "rel_drop": rel_drop}, related_upload_id=upload.id)
        anomalies_created += 1

    await db.commit()
    await db.refresh(upload)

    trust = await compute_trust(db)
    return {"upload": upload, "anomalies_created": anomalies_created, "idempotent": False, "trust": trust, "created_tasks": created_tasks, "updated_tasks": updated_tasks}
